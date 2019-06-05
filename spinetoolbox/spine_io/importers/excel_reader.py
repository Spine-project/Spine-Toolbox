# -*- coding: utf-8 -*-

from itertools import islice, takewhile
import io


from PySide2.QtWidgets import QFileDialog

from openpyxl import load_workbook

from spinedb_api import RelationshipClassMapping, ObjectClassMapping

from spine_io.io_api import SourceConnection


def select_excel_file(parent=None):
    """
    Launches QFileDialog with .txt filter
    """
    return QFileDialog.getOpenFileName(parent, "", "*.xlsx;*.xlsm;*.xltx;*.xltm")


class ExcelConnector(SourceConnection):
    """
    Template class to read data from another QThread
    """

    # name of data source, ex: "Text/CSV"
    DISPLAY_NAME = "Excel"

    # dict with option specification for source.
    OPTIONS = {
        "header": {'type': bool, 'label': 'Has header', 'default': False},
        "row": {'type': int, 'label': 'Skip rows', 'Minimum': 0, 'default': 0},
        "column": {'type': int, 'label': 'Skip columns', 'Minimum': 0, 'default': 0},
        "read_until_col": {'type': bool, 'label': 'Read until empty column on first row', 'default': False},
        "read_until_row": {'type': bool, 'label': 'Read until empty row on first column', 'default': False},
    }

    # Modal widget that that returns source object and action (OK, CANCEL)
    SELECT_SOURCE_UI = select_excel_file

    def __init__(self):
        super(ExcelConnector, self).__init__()
        self._filename = None
        self._wb = None

    def connect_to_source(self, source):
        """saves filepath
        
        Arguments:
            source {str} -- filepath
        """
        if source:
            self._filename = source
            try:
                # FIXME: there seems to be no way of closing the workbook
                # when read_only=True, read file into memory first and then
                # open to avoid locking file while toolbox is running.
                with open(self._filename, "rb") as bin_file:
                    in_mem_file = io.BytesIO(bin_file.read())
                self._wb = load_workbook(in_mem_file, read_only=True)
                return True
            except Exception as error:
                raise error

    def disconnect(self):
        """Disconnect from connected source.
        """
        if self._wb:
            self._wb.close()
        return True

    def get_tables(self):
        """Method that should return a list of table names, list(str)
        
        Raises:
            NotImplementedError: [description]
        """
        if not self._wb:
            return {}
        try:
            sheets = {}
            for sheet in self._wb.sheetnames:
                mapping, option = create_mapping_from_sheet(self._wb[sheet])
                sheets[sheet] = {"mapping": mapping, "options": option}
            return sheets
        except Exception as error:
            raise error

    def get_data_iterator(self, table, options, max_rows=-1):
        """
        Return data read from data source table in table. If max_rows is
        specified only that number of rows.
        """
        if not self._wb:
            return iter([]), [], 0

        if not table in self._wb:
            # table not found
            return iter([]), [], 0
        worksheet = self._wb[table]

        # get options
        has_header = options.get("header", False)
        skip_rows = options.get("row", 0)
        skip_columns = options.get("column", 0)
        stop_at_empty_col = options.get("read_until_col", False)
        stop_at_empty_row = options.get("read_until_row", False)

        if max_rows == -1:
            max_rows = None
        else:
            max_rows += skip_rows

        read_to_col = None
        try:
            first_row = next(islice(worksheet.iter_rows(), skip_rows, max_rows))
            if stop_at_empty_col:
                # find first empty col in top row and use that as a stop
                num_cols = 0
                for i, column in enumerate(islice(first_row, skip_columns, None)):
                    num_cols = i
                    if column.value is None:
                        read_to_col = i + skip_columns
                        break
            else:
                num_cols = len(first_row)
        except StopIteration:
            # no data
            num_cols = 0

        header = []
        rows = worksheet.iter_rows()
        rows = islice(rows, skip_rows, max_rows)
        # find header if it has one
        if has_header:
            try:
                header = [c.value for c in islice(next(rows), skip_columns, read_to_col)]
            except StopIteration:
                # no data
                return iter([]), [], 0

        # iterator for selected columns and and skipped rows
        data_iterator = (list(cell.value for cell in islice(row, skip_columns, read_to_col)) for row in rows)
        if stop_at_empty_row:
            # add condition to iterator
            condition = lambda row: row[0] is not None
            data_iterator = takewhile(condition, data_iterator)

        return data_iterator, header, num_cols


def create_mapping_from_sheet(worksheet):
    """
    Checks if sheet is a valid spine excel template, if so creates a
    mapping object for each sheet.
    """

    options = {"header": False, "row": 0, "column": 0, "read_until_col": False, "read_until_row": False}
    mapping = ObjectClassMapping()
    sheet_type = worksheet["A2"].value
    sheet_data = worksheet["B2"].value
    if not isinstance(sheet_type, str):
        return None, None
    if not isinstance(sheet_data, str):
        return None, None
    if sheet_type.lower() not in ["relationship", "object"]:
        return None, None
    if sheet_data.lower() not in ["parameter", "json array"]:
        return None, None
    if sheet_type.lower() == "relationship":
        mapping = RelationshipClassMapping()
        rel_dimension = worksheet["D2"].value
        rel_name = worksheet["C2"].value
        if not isinstance(rel_name, str):
            return None, None
        if not rel_name:
            return None, None
        if not isinstance(rel_dimension, int):
            return None, None
        if not rel_dimension >= 1:
            return None, None
        if sheet_data.lower() == "parameter":
            obj_classes = next(islice(worksheet.iter_rows(), 3, 4))
            obj_classes = [r.value for r in obj_classes[:rel_dimension]]
        else:
            obj_classes = islice(worksheet.iter_rows(), 3, 3 + rel_dimension)
            obj_classes = [r[0].value for r in obj_classes]
        if not all(isinstance(r, str) for r in obj_classes) or any(r is None or r.isspace() for r in obj_classes):
            return None, None
        if sheet_data.lower() == "parameter":
            options.update({"header": True, "row": 3, "read_until_col": True, "read_until_row": True})
            mapping = RelationshipClassMapping.from_dict(
                {
                    "map_type": "RelationshipClass",
                    "name": rel_name,
                    "object_classes": obj_classes,
                    "objects": list(range(rel_dimension)),
                    "parameters": {"map_type": "parameter", "name": {"map_type": "row", "value_reference": -1}},
                }
            )
        else:
            options.update({"header": False, "row": 3, "read_until_col": True, "read_until_row": False})
            mapping = RelationshipClassMapping.from_dict(
                {
                    "map_type": "RelationshipClass",
                    "name": rel_name,
                    "object_classes": obj_classes,
                    "objects": [{"map_type": "row", "value_reference": i} for i in range(rel_dimension)],
                    "parameters": {
                        "map_type": "parameter",
                        "name": {"map_type": "row", "value_reference": rel_dimension},
                        "extra_dimensions": [0],
                    },
                }
            )

    elif sheet_type.lower() == "object":
        obj_name = worksheet["C2"].value
        if not isinstance(obj_name, str):
            return None, None
        if not obj_name:
            return None, None
        if sheet_data.lower() == "parameter":
            options.update({"header": True, "row": 3, "read_until_col": True, "read_until_row": True})
            mapping = ObjectClassMapping.from_dict(
                {
                    "map_type": "ObjectClass",
                    "name": obj_name,
                    "object": 0,
                    "parameters": {"map_type": "parameter", "name": {"map_type": "row", "value_reference": -1}},
                }
            )
        else:
            options.update({"header": False, "row": 3, "read_until_col": True, "read_until_row": False})
            mapping = ObjectClassMapping.from_dict(
                {
                    "map_type": "ObjectClass",
                    "name": obj_name,
                    "object": {"map_type": "row", "value_reference": 0},
                    "parameters": {
                        "map_type": "parameter",
                        "name": {"map_type": "row", "value_reference": 1},
                        "extra_dimensions": [0],
                    },
                }
            )
    else:
        return None, None
    return mapping, options
